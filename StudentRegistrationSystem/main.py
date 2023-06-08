from tkinter import *
from datetime import date
from tkinter import filedialog
from tkinter import messagebox
from PIL import Image, ImageTk
import os
from tkinter.ttk import Combobox
import openpyxl, xlrd
from openpyxl import Workbook
import pathlib

background = "#06283D"
framebg = "#EDEDED"
framefg = "#06283D"

window = Tk()
window.title('Student Registration System')
window.geometry('1250x700+210+100')
window.resizable(width=False, height=False)
window.config(bg=background)

file = pathlib.Path('Student_data.xlsx')
if file.exists():
    pass
else:
    file = Workbook()
    sheet = file.active
    sheet['A1'] = "Registration No."
    sheet['B1'] = "Name"
    sheet['C1'] = "Class"
    sheet['D1'] = "Gender"
    sheet['E1'] = "DOB"
    sheet['F1'] = "Date of Registration"
    sheet['G1'] = "Religion"
    sheet['H1'] = "Skill"
    sheet['I1'] = "Father Name"
    sheet['J1'] = "Mother Name"
    sheet['K1'] = "Father's Occupation"
    sheet['L1'] = "Mother's Occupation"
    file.save('Student_data.xlsx')


def Exit():
    window.destroy()


def showimage():
    global filename
    global img
    filename = filedialog.askopenfile(initialdir=os.getcwd(),
                                      title="Select image file", filetypes=(("JPG File", "*.jpg"),
                                                                            ("PNG File", "*.png"),
                                                                            ("All files", "*.*")))
    if filename:
        img = Image.open(filename)
        resized_image = img.resize((190, 190))
        photo2 = ImageTk.PhotoImage(resized_image)
        lbl.config(image=photo2)
        lbl.image = photo2


def registration_no():
    file = openpyxl.load_workbook('Student_data.xlsx')
    workbook = openpyxl.load_workbook('Student_data.xlsx')
    worksheet = workbook.active

    max_rows = worksheet.max_row
    max_row_value = worksheet.cell(row=max_rows, column=1).value

    try:
        Registration.set(max_row_value + 1)
    except:
        Registration.set("1")


def Clear():
    global img
    Name.set('')
    DOB.set('')
    Religion.set('')
    Skill.set('')
    F_Name.set('')
    M_Name.set('')
    Father_Occupation.set('')
    Mother_Occupation.set('')
    Class.set("Select Class")
    registration_no()


    img = None
    img1 = Image.open("images/siluette2.png")
    resized_image = img1.resize((190, 190))
    photo2 = ImageTk.PhotoImage(resized_image)
    lbl.config(image=photo2)
    lbl.image = photo2



def Save():
    R1 = Registration.get()
    N1 = Name.get()
    C1 = Class.get()
    try:
        G1 = gender
    except:
        messagebox.showerror("Error", "Select Gender")
    D2 = DOB.get()
    D1 = Date.get()
    Re1 = Religion.get()
    S1 = Skill.get()
    fathername = F_Name.get()
    mothername = M_Name.get()
    F1 = Father_Occupation.get()
    M1 = Mother_Occupation.get()

    if N1 == "" or G1 == "Select Class" or D2 == "" or Re1 == "" or S1 == "" or fathername == "" or mothername == "" or F1 == "" or M1 == "":
        messagebox.showerror("error", "Few data is missing!")
    else:
        file = openpyxl.load_workbook('Student_data.xlsx')
        sheet = file.active
        sheet.cell(column=1, row=sheet.max_row + 1, value=R1)
        sheet.cell(column=2, row=sheet.max_row, value=N1)
        sheet.cell(column=3, row=sheet.max_row, value=C1)
        sheet.cell(column=4, row=sheet.max_row, value=G1)
        sheet.cell(column=5, row=sheet.max_row, value=D2)
        sheet.cell(column=6, row=sheet.max_row, value=D1)
        sheet.cell(column=7, row=sheet.max_row, value=Re1)
        sheet.cell(column=8, row=sheet.max_row, value=S1)
        sheet.cell(column=9, row=sheet.max_row, value=fathername)
        sheet.cell(column=10, row=sheet.max_row, value=mothername)
        sheet.cell(column=11, row=sheet.max_row, value=F1)
        sheet.cell(column=12, row=sheet.max_row, value=M1)

        file.save(r'Student_data.xlsx')

        try:
            img.save("Student Images/" + str(R1) + ".jpg")
        except:
            messagebox.showinfo("info", "Profile Picture is not available!!!!!")

        messagebox.showinfo("info", "sucessfully data entered!!!")

        Clear()

        registration_no()


def searchbutton():
    text = search.get()
    Clear()
    #SaveButton.config(state='disable')

    file = openpyxl.load_workbook("Student_data.xlsx")
    sheet = file.active

    for row in sheet.rows:
        if row[0].value == int(text):
            name = row[0]
            reg_no_position = str(name)[14:-1]
            reg_number = str(name)[15:-1]
            try:
                img = Image.open("Student Images/" + reg_number + ".jpg")
                resized_image = img.resize((190, 190))
                photo2 = ImageTk.PhotoImage(resized_image)
                lbl.config(image=photo2)
                lbl.image = photo2
            except:
                messagebox.showinfo("info", "Profile Picture is not available!!!!!")

            Registration.set(reg_number)
            Name.set(row[1].value)
            Class.set(row[2].value)
            gender = row[3].value
            DOB.set(row[4].value)
            Date.set(row[5].value)
            Religion.set(row[6].value)
            Skill.set(row[7].value)
            F_Name.set(row[8].value)
            M_Name.set(row[9].value)
            Father_Occupation.set(row[10].value)
            Mother_Occupation.set(row[11].value)
            #SaveButton.config(state='disable')

def Update():
    R1 = Registration.get()
    N1 = Name.get()
    C1 = Class.get()
    selection()
    G1 = gender
    D2 = DOB.get()
    D1 = Date.get()
    Re1 = Religion.get()
    S1 = Skills.get()
    fathername = F_Name.get()
    mothername = M_Name.get()
    F1 = Father_Occupation.get()
    M1 = Mother_Occupation.get()

    file = openpyxl.load_workbook("Student_data.xlsx")
    sheet = file.active

    for row in sheet.rows:
        if row[0].value == R1:
            name = row[0]
            reg_no_position = str(name)[14:-1]
            reg_number = str(name)[15:-1]
    sheet.cell(column=1, row=int(reg_number), value=R1)
    sheet.cell(column=2, row=int(reg_number), value=N1)
    sheet.cell(column=3, row=int(reg_number), value=C1)
    sheet.cell(column=4, row=int(reg_number), value=G1)
    sheet.cell(column=5, row=int(reg_number), value=D2)
    sheet.cell(column=6, row=int(reg_number), value=D1)
    sheet.cell(column=7, row=int(reg_number), value=Re1)
    sheet.cell(column=8, row=int(reg_number), value=S1)
    sheet.cell(column=9, row=int(reg_number), value=fathername)
    sheet.cell(column=10, row=int(reg_number), value=mothername)
    sheet.cell(column=11, row=int(reg_number), value=F1)
    sheet.cell(column=12, row=int(reg_number), value=M1)
    file.save(r'Student_data.xlsx')
    try:
        img.save("student images/" + str(R1) + ".jpg")
    except:
        pass

    messagebox.showinfo("Update", "Update Sucessfully!!")

    Clear()

def selection():
    global gender
    value = radio.get()
    if value == 1:
        gender = "Male"
    elif value == 2:
        gender = "Female"
    else:
        gender = "other"


Label(window, text="Email: georgethomidis@gmail.com", width=10, height=3, bg="#f0687c", anchor='e').pack(side=TOP,fill=X)
Label(window, text="STUDENT REGISTRATION", width=10, height=2, bg="#c36464", fg="#fff", font="arial 20 bold").pack(side=TOP, fill=X)

search = StringVar()
Entry(window, textvariable=search, width=15, bd=2, font="arial 20").place(x=820, y=70)
imageicon3 = PhotoImage(file="images/149852.png")
Srch = Button(window, text="Search", compound=LEFT, image=imageicon3, width=123, bg='#68ddfa', font="arial 13 bold",command=searchbutton)
Srch.place(x=1060, y=72)

imageicon4 = PhotoImage(file="images/layer.png")
update_button = Button(window, image=imageicon4, bg="#c36464")
update_button.place(x=110, y=64)

Label(window, text="Registration No:", font="arial 13", fg=framebg, bg=background).place(x=30, y=150)
Label(window, text="Date:", font="arial 13", fg=framebg, bg=background).place(x=500, y=150)

Registration = IntVar()
Date = StringVar()

reg_entry = Entry(window, textvariable=Registration, width=15, font="arial 10")
reg_entry.place(x=160, y=150)

registration_no()

today = date.today()
d1 = today.strftime("%d/%m/%Y")
date_entry = Entry(window, textvariable=Date, width=15, font="arial 10")
date_entry.place(x=550, y=150)

Date.set(d1)

obj1 = LabelFrame(window, text="Student's Details", font=20, bd=2, width=900, bg=framebg, fg=framefg, height=250,relief=GROOVE)
obj1.place(x=30, y=200)

Label(obj1, text="Full name:", font="arial 13", bg=framebg, fg=framefg).place(x=30, y=50)
Label(obj1, text="Date of Birth:", font="arial 13", bg=framebg, fg=framefg).place(x=30, y=100)
Label(obj1, text="Gender:", font="arial 13", bg=framebg, fg=framefg).place(x=30, y=150)

Label(obj1, text="Class:", font="arial 13", bg=framebg, fg=framefg).place(x=500, y=50)
Label(obj1, text="Religion:", font="arial 13", bg=framebg, fg=framefg).place(x=500, y=100)
Label(obj1, text="Skills:", font="arial 13", bg=framebg, fg=framefg).place(x=500, y=150)

Name = StringVar()
name_entry = Entry(obj1, textvariable=Name, width=20, font="aria 10")
name_entry.place(x=160, y=50)

DOB = StringVar()
dob_entry = Entry(obj1, textvariable=DOB, width=20, font="aria 10")
dob_entry.place(x=160, y=100)

radio = IntVar()
R1 = Radiobutton(obj1, text="Male", variable=radio, value=1, bg=framebg, fg=framefg, command=selection)
R1.place(x=150, y=150)

R2 = Radiobutton(obj1, text="Female", variable=radio, value=2, bg=framebg, fg=framefg, command=selection)
R2.place(x=200, y=150)

R3 = Radiobutton(obj1, text="other", variable=radio, value=3, bg=framebg, fg=framefg, command=selection)
R3.place(x=265, y=150)

Religion = StringVar()
religion_entry = Entry(obj1, textvariable=Religion, width=20, font="aria 10")
religion_entry.place(x=630, y=100)

Skill = StringVar()
skill_entry = Entry(obj1, textvariable=Skill, width=20, font="aria 10")
skill_entry.place(x=630, y=150)

Class = Combobox(obj1, values=['1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12'], font="Roboto 10",width=17, state="r")
Class.place(x=630, y=50)
Class.set("Select Class")

obj2 = LabelFrame(window, text="Parent's Details", font=20, bd=2, width=900, bg=framebg, fg=framefg, height=220,relief=GROOVE)
obj2.place(x=30, y=470)

Label(obj2, text="Father's Name:", font="arial 13", bg=framebg, fg=framefg).place(x=30, y=50)
Label(obj2, text="Occupation:", font="arial 13", bg=framebg, fg=framefg).place(x=30, y=100)

F_Name = StringVar()
f_entry = Entry(obj2, textvariable=F_Name, width=20, font="arial 10")
f_entry.place(x=160, y=50)

Father_Occupation = StringVar()
fo_entry = Entry(obj2, textvariable=Father_Occupation, width=20, font="arial 10")
fo_entry.place(x=160, y=100)

Label(obj2, text="Mother's Name:", font="arial 13", bg=framebg, fg=framefg).place(x=500, y=50)
Label(obj2, text="Occupation:", font="arial 13", bg=framebg, fg=framefg).place(x=500, y=100)

M_Name = StringVar()
m_entry = Entry(obj2, textvariable=M_Name, width=20, font="arial 10")
m_entry.place(x=630, y=50)

Mother_Occupation = StringVar()
mo_entry = Entry(obj2, textvariable=Mother_Occupation, width=20, font="arial 10")
mo_entry.place(x=630, y=100)

f = Frame(window, bd=3, bg="black", width=200, height=200, relief=GROOVE)
f.place(x=1000, y=150)

img = PhotoImage(file="images/siluette2.png")
lbl = Label(f, bg="black", image=img)
lbl.place(x=0, y=0)

Button(window, text="Upload", width=19, height=2, font="arial 12 bold", bg="lightblue", command=showimage).place(x=1000,y=370)

SaveButton = Button(window, text="Save", width=19, height=2, font="arial 12 bold", bg="lightgreen", command=Save).place(x=1000,y=450)

Button(window, text="Reset", width=19, height=2, font="arial 12 bold", bg="lightpink", command=Clear).place(x=1000,y=530)

Button(window, text="Exit", width=19, height=2, font="arial 12 bold", bg="grey", command=Exit).place(x=1000, y=610)

window.mainloop()
